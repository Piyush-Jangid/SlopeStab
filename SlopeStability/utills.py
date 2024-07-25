import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.cm import ScalarMappable
import pandas as pd
import numpy as np
import math 
from openpyxl import load_workbook
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
import pkg_resources
import os

def getDataFromExcel(file_name,sheet_name):
    # DATA_FOLDER = os.path.join(os.path.dirname(__file__), 'Data')
    
    # file_path = os.path.join(DATA_FOLDER, file_name)
    file_path = pkg_resources.resource_filename('SlopeStability', 'Data')
    file_path = os.path.join(file_path,file_name)
    # print(file_path)
    wb = load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2,values_only=True):
            # Filter out rows that are None or do not have at least 2 valid values
            if len(row) >= 2 and row[0] is not None and row[1] is not None:
                data.append([row[0], row[1]])  # Select only the first 2 columns

    # Convert the data to a numpy array
    data = np.array(data)
    return data
    pass

def find_closest_keys(d, keys):
    # This function finds the two keys closest to 'd' (one higher, one lower)
    arr = np.array(sorted(keys))
    low, high=-1,-1
    n = len(arr)
    
    for i in range(n-1):
        if d>arr[i] and d<arr[i+1]:
            low = (arr[i])
            high = (arr[i+1])
    if low==-1 and high==-1:
        raise ValueError("Point is out of bounds of the Stability Charts")
    return low,high
    pass

def linear_interpolate(x1, y1, x2, y2, x):
    # Perform linear interpolation
    if x2 - x1 == 0:  # To handle division by zero if x1 == x2
        return y1
    return y1 + (y2 - y1) * (x - x1) / (x2 - x1)
    pass

def interpolate_2d_array(points, x):
    # Ensure points are sorted by x values
    points = np.array(sorted(points, key=lambda point: point[0]))
    
    # Extract x and y arrays from the points
    x_points = points[:, 0]
    y_points = points[:, 1]
    
    # Check if x is out of the bounds of the provided points
    if x < x_points[0] or x > x_points[-1]:
        raise ValueError("Point is out of bounds of the Stability Charts.")
    
    # Find the indices of the two x points between which x lies
    for i in range(len(x_points) - 1):
        if x_points[i] <= x <= x_points[i + 1]:
            x0, y0 = x_points[i], y_points[i]
            x1, y1 = x_points[i + 1], y_points[i + 1]
            # Apply linear interpolation formula
            y = y0 + (x - x0) * (y1 - y0) / (x1 - x0)
            return y
    n = len(x_points)
    if math.isclose(x_points[n-1], x, rel_tol=1e-8, abs_tol=1e-8):
        return y_points[n-1]
    
    # If we can't find the interval, raise an error
    raise ValueError("Interpolation could not be performed. Check input data.")
    pass

def getFromDict(m,d,x):
    if d in m:
        result = interpolate_2d_array(m[d], x)
    else:
        # Find the closest keys for interpolation
        lower_key, upper_key = find_closest_keys(d, m.keys())
        # Calculate f for both lower and upper keys
        lower_value = interpolate_2d_array(m[lower_key], x)
        upper_value = interpolate_2d_array(m[upper_key], x)
        # Interpolate to find the result
        result = linear_interpolate(lower_key, lower_value, upper_key, upper_value, d)

    return result
    pass

def FailureCircle(x0,y0,D,H,T,clr,ax):
    # Define the circle's center and radius
    center = (x0, y0)
    if T==1:
        radius = np.sqrt(x0*x0+y0*y0)
    else:
        radius = y0+D\
    # Create an array of angles from 0 to 2pi
    theta = np.linspace(0, 2 * np.pi, 2000)

    # Parametric equations for the circle
    x = center[0] + radius * np.cos(theta)
    y = center[1] + radius * np.sin(theta)

    # Filter the circle points to show only where y <= 5 and x >= 0
    x1 = []
    y1 = []

    # Iterate through the arrays and apply the conditions
    for i in range(len(x)):
        if (x[i] >= 0 and y[i] <= H) or (y[i]<=0 and x[i]<=0) :
            x1.append(x[i])
            y1.append(y[i])
    
    # Plot the filtered circle
    ax.plot(x1, y1, color=clr)
    pass

def DrawSolution(x0_list,y0_list,D_list,H,beta,T_list,q,Hw,Hwdash,fos_values):
    median_fos = np.median(fos_values)
    x0_list = np.array(x0_list)
    y0_list = np.array(y0_list)
    D_list = np.array(D_list)
    T_list = np.array(T_list)
    
    # Create a figure and an axes
    fig, ax = plt.subplots(figsize=(12, 20))
    
    D = max(D_list)
    
    if D==0:
        D=H
        
    # Convert slope angle from degrees to radians
    beta_rad=np.radians(beta)
    # Plot the slope line
    line_length = H / np.sin(beta_rad)
    line_x = [0, line_length * np.cos(beta_rad)]
    line_y = [0, H]

    # Ensure only positive x values (i.e., the line should point to the right)
    if line_x[1] < 0:
        line_x[1] = -line_x[1]
        line_y[1] = -line_y[1]

    # Plot the slope line
    ax.plot(line_x, line_y, label='Slope Line', color='#C04000')
    
    r1=np.sqrt((x0_list)**2+(y0_list)**2)
    r2=(y0_list)+(D_list)
    a = max(x0_list) + max(max(r1),max(r2))+7
    b = max(max(np.sqrt(r1**2-y0_list**2)),max(np.sqrt(r2**2-y0_list**2)))+5
    
    # Plot the horizontal line at the height of the slope
    x_values = np.linspace(H / np.tan(beta_rad), a, 40)  # Adjust max value if needed
    y_values = np.full_like(x_values, H)  # y=H for all x in x_values
    ax.plot(x_values, y_values, color='#C04000', label='Horizontal Line at H')


# Plot the horizontal line at the depth of the foundation
    x_values_depth = np.linspace(-b, 0, 40)
    y_values_depth = np.full_like(x_values_depth, 0)  # y=0 for all x in x_values_depth
    ax.plot(x_values_depth, y_values_depth, color='#C04000', label='Foundation Depth Line')
    
    x_values_depth = np.linspace(-b,a , 80)
    y_values_depth = np.full_like(x_values_depth, -D)  # y=0 for all x in x_values_depth
    ax.plot(x_values_depth, y_values_depth, color='#C04000', label='Foundation Depth Line')
    
    x_values_depth = np.linspace(0, a, 40)
    y_values_depth = np.full_like(x_values_depth,0)  # y=0 for all x in x_values_depth
    ax.plot(x_values_depth, y_values_depth, linestyle='--',color='red', label='Foundation Depth Line')

    x_const = -b
    y_range = np.linspace(0,-D, 40)
    # Plot the line parallel to y-axis
    ax.plot([x_const]*len(y_range), y_range,color='#C04000')
    
    x_const = a
    y_range = np.linspace(-D,H, 40)
    # Plot the line parallel to y-axis
    ax.plot([x_const]*len(y_range), y_range,color='#C04000')
    
    if Hwdash!=0:
        x_values_depth = np.linspace(Hwdash/np.tan(beta_rad),a, 40)
        y_values_depth = np.full_like(x_values_depth, Hwdash)  # y=0 for all x in x_values_depth
        ax.plot(x_values_depth, y_values_depth, color='blue', label='Foundation Depth Line')
        
        ax.annotate(f'Hwdash={Hwdash} feet/m', xy=(a/2.5, Hwdash), xytext=(a/2.5,Hwdash+4),arrowprops=dict(facecolor='blue', shrink=0.1))

    # Annotate the slope angle
    ax.annotate(f'Angle: {beta}°', xy=(line_x[1]/2, line_y[1]/2), xytext=(line_x[1]/2 + 5, line_y[1]/2),
                arrowprops=dict(facecolor='black', arrowstyle='->'), fontsize=12, color='black')
    
    ax.annotate(f'H: {H} feet/m', xy=(a, H/2), xytext=(a+2, H/2),
                arrowprops=dict(facecolor='black', arrowstyle='->'), fontsize=12, color='black')
    
    if D!=0:
        ax.annotate(f'D: {D} feet/m', xy=(-b, -D/2), xytext=(-(b+14.5), -D/2),
            arrowprops=dict(facecolor='black', arrowstyle='->'), fontsize=12, color='black')
    if Hw!=0:
        x_values_depth = np.linspace(-(b), Hw/np.tan(beta_rad), 40)
        y_values_depth = np.full_like(x_values_depth, Hw)  # y=0 for all x in x_values_depth
        ax.plot(x_values_depth, y_values_depth, color='blue', label='Foundation Depth Line')
        x_const = -b
        y_range = np.linspace(0,Hw, 40)
        # Plot the line parallel to y-axis
        ax.plot([x_const]*len(y_range), y_range,color='blue')
        ax.annotate('', xy=(-b/2, Hw), xytext=(-b/2,Hw+1),arrowprops=dict(facecolor='blue', shrink=0.1))
        ax.annotate(f'Hw: {Hw} feet/m', xy=(-b, Hw/2), xytext=(-(b+14.5), Hw/2),
                arrowprops=dict(facecolor='black', arrowstyle='->'), fontsize=12, color='black')
    if q!=0:
        num_arrows = 5  # Number of arrows
        x_arrows = np.linspace(H / np.tan(beta_rad),a, num_arrows)

        # Length of each arrow
        arrow_length = 2  # Shorter length
        i=0
        for x in x_arrows:
            if(i==np.round(len(x_arrows)/2)):
                ax.annotate(f'q={q} kPa/psf', xy=(x, H), xytext=(x,H+2.5),arrowprops=dict(facecolor='black',shrink=0.1))  # Adjust shrink for smaller arrow
            else:
                ax.annotate('', xy=(x, H), xytext=(x,H+arrow_length),arrowprops=dict(facecolor='black',shrink=0.1))
            i=i+1

    # Set the aspect of the plot to be equal
    n = len(x0_list)
    cmap = mcolors.LinearSegmentedColormap.from_list('yellow_to_red', ['yellow', 'red'])
    if n>1:
        # Generate the colors from the colormap
        colors = [mcolors.rgb2hex(cmap(i / (n - 1))) for i in range(n)]
        # print(colors)
        for i in range(n):
            clr = colors[i]
            FailureCircle(x0_list[i], y0_list[i], D_list[i], H, T_list[i], clr, ax)
    else:
        FailureCircle(x0_list[0],y0_list[0],D_list[0],H,T_list[0],'yellow',ax)

    
    x0_list, y0_list, D_list, T_list = map(np.array, (x0_list, y0_list, D_list, T_list))
    median_index = np.argsort(fos_values)[len(fos_values) // 2]
    median_fos = fos_values[median_index]
    median_x0 = x0_list[median_index]
    median_y0 = y0_list[median_index]
    median_D = D_list[median_index]
    median_T = T_list[median_index]

    FailureCircle(median_x0,median_y0, median_D, H, median_T, 'black', ax)

    # Adding color bar at top-left corner
    norm = plt.Normalize(vmin=min(fos_values), vmax=max(fos_values))
    sm = ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])  # Dummy array for the ScalarMappable
    cbar_ax = inset_axes(ax, width="50%", height="70%", loc='upper left', bbox_to_anchor=(0.05, 0.95, 0.9, 0.1), bbox_transform=ax.transAxes)
    cbar = plt.colorbar(sm, cax=cbar_ax, orientation='horizontal') #(0.1,0.95,1,1)
    cbar.set_label('Factor of Safety', labelpad=-60,size=15)
    
    ax.set_aspect('equal')
    # Set labels and title
    ax.set_xlabel('x')
    ax.set_ylabel('y')
    # ax.set_title('Solution',pad=30,size=20)
    ax.axis('off')

    # plt.show() 
    pass

def generate_samples(mean, cov, dist_type, num_samples):
    if dist_type == 'normal':
        std = mean * cov
        samples = np.random.normal(mean, std, num_samples)
        # Ensure samples are positive
        samples = np.clip(samples, 0, None)
    elif dist_type == 'lognormal':
        std = mean * cov
        mean_ln = np.log(mean**2 / np.sqrt(std**2 + mean**2))
        sigma_ln = np.sqrt(np.log(std**2 / mean**2 + 1))
        samples = np.random.lognormal(mean_ln, sigma_ln, num_samples)
    elif dist_type == 'uniform':
        std = mean * cov
        lower = mean - std * np.sqrt(3)
        upper = mean + std * np.sqrt(3)
        samples = np.random.uniform(lower, upper, num_samples)
        # Ensure samples are positive
        samples = np.clip(samples, 0, None)
    else:
        raise ValueError(f"Unsupported distribution type: {dist_type} Supported only normal, lognormal and uniform")
    return samples
    pass